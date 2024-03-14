import inspect
import logging
import openpyxl
import string
import random
import os
from colorlog import ColoredFormatter
from openpyxl.styles import PatternFill
class utills():
    def custom_logger(logLevel=logging.DEBUG, log_folder="Logs"):
        # set class /method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        log_file_path = os.path.join(log_folder, "ArtemusH.log")
        fh= logging.FileHandler(log_file_path)
        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add consile handler to logger
        logger.addHandler(fh)
        return logger


    def custom_logger_5106(logLevel=logging.DEBUG, log_folder="Logs"):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        log_file_path = os.path.join(log_folder, "Artemus5106.log")
        fh= logging.FileHandler(log_file_path)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)

    def readData(file, sheetName, rownum, coloumnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rownum, coloumnno).value

    def writeData(file, sheetName, rownum, coloumnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rownum, coloumnno).value = data
        workbook.save(file)

    def random_invoceGenerator(size=8, chars=string.ascii_uppercase + string.digits):
        return ''.join(random.choice(chars) for x in range(size))

    def random_BillGenerator(size=6, chars=string.ascii_uppercase + string.digits):
        return ''.join(random.choice(chars) for x in range(size))

    def random_importerNameGenerator(size=4, chars=string.ascii_uppercase):
        return ''.join(random.choice(chars) for x in range(size))

    def random_importerNumberGenerator(size=9, chars=string.digits):
        return ''.join(random.choice(chars) for x in range(size))


